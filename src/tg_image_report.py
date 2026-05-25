"""
NSE Telegram + WhatsApp Image Report  — Single Combined Image
=============================================================
Generates ONE image with 4 sections in a 2x2 grid:
  Top-left:     Bullish Top 10  (Symbol · Close · Chg% · Vol)
  Top-right:    Bearish Top 10  (Symbol · Close · Chg% · Vol)
  Bottom-left:  Neutral Top 10  (Symbol · Close · Chg% · Vol)
  Bottom-right: Most Active Top 10 (Symbol · Close · Chg% · Vol)

Sends the single image to Telegram + WhatsApp.
Canvas height is computed exactly from content — no empty space.
"""

import os, sys, glob, re, base64
import requests
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch
import pandas as pd
import numpy as np
from datetime import datetime
import pytz

# ── Config ────────────────────────────────────────────────────────────────────
TG_BOT_TOKEN  = os.environ.get("TG_BOT_TOKEN", "")
TG_CHAT_ID    = os.environ.get("TG_CHAT_ID", "")
OUTPUT_DIR    = os.environ.get("OUTPUT_DIR", "output_data")
INDICATOR_DIR = os.environ.get("INDICATOR_DIR", "indicator_data")
MAPPING_FILE  = "src/mc_nse_mapping.csv"

# WhatsApp via Twilio + ImgBB
TWILIO_SID    = os.environ.get("TWILIO_SID", "")
TWILIO_TOKEN  = os.environ.get("TWILIO_AUTH_TOKEN", "")
_wa_from_raw  = os.environ.get("WA_FROM_NUMBER", "")
WA_FROM       = _wa_from_raw if _wa_from_raw.startswith("whatsapp:") else f"whatsapp:{_wa_from_raw}" if _wa_from_raw else ""
_wa_to_raw    = os.environ.get("WA_TO_NUMBER", "+97450740794")
WA_TO         = _wa_to_raw if _wa_to_raw.startswith("whatsapp:") else f"whatsapp:{_wa_to_raw}"
IMGBB_KEY     = os.environ.get("IMGBB_API_KEY", "")

IST   = pytz.timezone("Asia/Kolkata")
NOW   = datetime.now(IST)
TOP_N = 10

# ── Colours ───────────────────────────────────────────────────────────────────
WHITE  = "#FFFFFF"
LGREY  = "#F4F4F4"
MGREY  = "#DDDDDD"
BGPAGE = "#F0F2F5"
DARK   = "#1A1A1A"
MUTED  = "#888888"
GREEN  = "#1B8C3E"
RED    = "#C0392B"
BLUE   = "#2471A3"
ORANGE = "#D4680A"
HEADER = "#0D1B2A"


# ── Load latest output report ─────────────────────────────────────────────────
def load_report():
    def _date_key(f):
        m = re.search(r"(\d{2}-\w{3}-\d{4})", f)
        if m:
            try: return datetime.strptime(m.group(1), "%d-%b-%Y")
            except: pass
        return datetime.min

    files = sorted(glob.glob(f"{OUTPUT_DIR}/NSE_Indicator_Report_*.csv"),
                   key=_date_key, reverse=True)
    if not files:
        print("[ERROR] No report CSV found in", OUTPUT_DIR)
        sys.exit(1)
    path = files[0]
    print(f"[DATA] Loading: {path}")
    df = pd.read_csv(path, dtype=str)
    print(f"[DATA] Total rows: {len(df):,}")

    for col in ["CLOSE", "PREV_CLOSE", "LTP_MC", "VOLUME"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Compute CHANGE_PCT from bhavcopy prices
    if "CLOSE" in df.columns and "PREV_CLOSE" in df.columns:
        df["CHANGE_PCT"] = (
            (df["CLOSE"] - df["PREV_CLOSE"]) / df["PREV_CLOSE"].replace(0, np.nan)
        ) * 100
    elif "LTP_MC" in df.columns and "PREV_CLOSE" in df.columns:
        df["CHANGE_PCT"] = (
            (df["LTP_MC"] - df["PREV_CLOSE"]) / df["PREV_CLOSE"].replace(0, np.nan)
        ) * 100
    else:
        df["CHANGE_PCT"] = np.nan

    if "CLOSE" in df.columns and "LTP_MC" in df.columns:
        df["CLOSE"] = df["CLOSE"].fillna(df["LTP_MC"])
    if "PREV_CLOSE" not in df.columns:
        df["PREV_CLOSE"] = np.nan
    if "VOLUME" not in df.columns:
        df["VOLUME"] = np.nan

    if "SERIES" in df.columns:
        df = df[df["SERIES"].fillna("").str.upper() == "EQ"]
    if "NSE_SYMBOL" in df.columns:
        df = df[df["NSE_SYMBOL"].notna() & (df["NSE_SYMBOL"].str.strip() != "")]

    m = re.search(r"(\d{2}-\w{3}-\d{4})", path)
    report_date = m.group(1) if m else NOW.strftime("%d-%b-%Y")
    print(f"[DATA] EQ rows: {len(df):,} | Date: {report_date}")
    return df, report_date


# ── Build top-10 per trend ────────────────────────────────────────────────────
def top10_trend(df, trend):
    sub = df[df["Trend"].fillna("").str.lower() == trend.lower()].copy()
    if sub.empty:
        return pd.DataFrame()

    counts = (
        sub.groupby("NSE_SYMBOL").size()
        .reset_index(name="Count")
        .sort_values("Count", ascending=False)
        .head(TOP_N)
        .reset_index(drop=True)
    )
    top_syms = counts["NSE_SYMBOL"].tolist()
    sub_top  = sub[sub["NSE_SYMBOL"].isin(top_syms)]

    price_cols = [c for c in ["CLOSE", "PREV_CLOSE", "CHANGE_PCT", "VOLUME"]
                  if c in sub_top.columns]
    if price_cols:
        price = sub_top.groupby("NSE_SYMBOL")[price_cols].median().reset_index()
    else:
        price = pd.DataFrame({"NSE_SYMBOL": top_syms})

    result = counts.merge(price, on="NSE_SYMBOL", how="left")
    for col in ["CLOSE", "PREV_CLOSE", "CHANGE_PCT", "VOLUME"]:
        if col not in result.columns:
            result[col] = np.nan
    return result


# ── MC mapping + movers ───────────────────────────────────────────────────────
def load_mc_mapping():
    if os.path.exists(MAPPING_FILE):
        try:
            mc_df = pd.read_csv(MAPPING_FILE, dtype=str)
            if "MC_Code" in mc_df.columns and "NSE_Symbol" in mc_df.columns:
                return dict(zip(mc_df["MC_Code"].fillna(""), mc_df["NSE_Symbol"].fillna("")))
        except Exception:
            pass
    return {}


def load_movers(report_date):
    folders = sorted([f for f in glob.glob(f"{INDICATOR_DIR}/*") if os.path.isdir(f)],
                     reverse=True)
    date_folder = next((f for f in folders if os.path.basename(f) == report_date), None)
    if not date_folder and folders:
        date_folder = folders[0]
        print(f"[MOVERS] Using latest folder: {date_folder}")

    out = {}
    for fname in ["gainers.csv", "losers.csv", "most_active.csv"]:
        path = os.path.join(date_folder, fname) if date_folder else ""
        if path and os.path.exists(path):
            try:
                out[fname] = pd.read_csv(path, dtype=str)
            except Exception:
                out[fname] = pd.DataFrame()
        else:
            out[fname] = pd.DataFrame()
    return out


def prep_mover_df(df, mc_map):
    if df.empty:
        return df
    df = df.copy()

    if "nsCode" in df.columns:
        df["NSE_SYMBOL"] = df["nsCode"].fillna("")
    elif "MC_Code" in df.columns:
        df["NSE_SYMBOL"] = df["MC_Code"].map(mc_map).fillna(
            df.get("shortName", df.get("StockName", "")))
    elif "shortName" in df.columns:
        df["NSE_SYMBOL"] = df["shortName"]
    else:
        df["NSE_SYMBOL"] = ""

    col_map = [
        ("currPrice",      "CLOSE"),
        ("lastPrice",      "CLOSE"),
        ("prevPrice",      "PREV_CLOSE"),
        ("previousClose",  "PREV_CLOSE"),
        ("pChange",        "CHANGE_PCT"),
        ("volume",         "VOLUME"),
        ("quantityTraded", "VOLUME"),
    ]
    for src, dst in col_map:
        if src in df.columns and dst not in df.columns:
            df[dst] = df[src]

    for col in ["CLOSE", "PREV_CLOSE", "CHANGE_PCT", "VOLUME"]:
        df[col] = pd.to_numeric(df.get(col, np.nan), errors="coerce")

    # Compute CHANGE_PCT from CLOSE/PREV_CLOSE when pChange not in source (e.g. most_active.csv)
    mask = df["CHANGE_PCT"].isna() & df["CLOSE"].notna() & df["PREV_CLOSE"].notna() & (df["PREV_CLOSE"] != 0)
    df.loc[mask, "CHANGE_PCT"] = ((df.loc[mask, "CLOSE"] - df.loc[mask, "PREV_CLOSE"]) / df.loc[mask, "PREV_CLOSE"]) * 100

    df = df[df["NSE_SYMBOL"].astype(str).str.strip() != ""].reset_index(drop=True)
    return df.head(TOP_N)


# ── Drawing helpers ───────────────────────────────────────────────────────────
def rrect(ax, x, y, w, h, fc, ec="none", lw=0, alpha=1, r=0.05, zorder=1):
    ax.add_patch(FancyBboxPatch((x, y), w, h,
        boxstyle=f"round,pad={r}", linewidth=lw,
        facecolor=fc, edgecolor=ec, alpha=alpha, zorder=zorder))


def circle(ax, cx, cy, r, fc, zorder=2):
    ax.add_patch(plt.Circle((cx, cy), r, color=fc, zorder=zorder))


def fmt_vol(v):
    """Format volume: >=1Cr -> xCr, else xL"""
    try:
        v = float(v)
        if np.isnan(v):
            return "-"
        if v >= 1e7:
            return f"{v/1e7:.1f}Cr"
        return f"{v/1e5:.1f}L"
    except Exception:
        return "-"


# ── Draw one section (Bullish / Bearish / Neutral / Most Active) ──────────────
# All sections: # | SYMBOL | CLOSE | CHG% | VOL
def draw_section(ax, x, y, w, title, accent, icon, df_rows,
                 SEC_HDR, COL_H, ROW_H):
    col = {
        "rk":  x + 0.17,
        "sym": x + 0.36,
        "cls": x + w - 1.55,
        "chg": x + w - 0.90,
        "vol": x + w - 0.22,
    }

    # Section header bar
    rrect(ax, x, y - SEC_HDR, w, SEC_HDR, accent, r=0.04, zorder=3)
    ax.text(x + 0.26, y - SEC_HDR/2, icon,
            ha="center", va="center", fontsize=9, color=WHITE,
            fontweight="bold", zorder=5)
    ax.text(x + 0.42, y - SEC_HDR/2, title,
            ha="left", va="center", fontsize=6.5, color=WHITE,
            fontweight="bold", zorder=5)
    y -= SEC_HDR

    # Column header row
    rrect(ax, x, y - COL_H, w, COL_H, accent, alpha=0.12, r=0.03, zorder=2)
    cy = y - COL_H / 2
    hs = dict(ha="center", va="center", fontsize=4.8,
              color=accent, fontweight="bold", zorder=3)
    ax.text(col["rk"],       cy, "#",      **hs)
    ax.text(col["sym"]+0.04, cy, "SYMBOL", ha="left", va="center",
            fontsize=4.8, color=accent, fontweight="bold", zorder=3)
    ax.text(col["cls"], cy, "CLOSE",  **hs)
    ax.text(col["chg"], cy, "CHG%",   **hs)
    ax.text(col["vol"], cy, "VOL",    **hs)
    y -= COL_H

    # Data rows
    rows = df_rows.itertuples(index=False) if hasattr(df_rows, "itertuples") else iter(df_rows)
    for i, row in enumerate(rows):
        if i >= TOP_N:
            break

        # Support both DataFrame rows (namedtuple) and plain tuples
        if hasattr(row, "NSE_SYMBOL"):
            sym     = str(row.NSE_SYMBOL)
            cls_v   = float(row.CLOSE)      if not pd.isna(row.CLOSE)      else np.nan
            chg_v   = float(row.CHANGE_PCT) if not pd.isna(row.CHANGE_PCT) else np.nan
            vol_raw = row.VOLUME
        else:
            sym, cls_v, _, chg_v, vol_raw = row[0], row[1], row[2], row[3], row[4]

        try: cls_v = float(cls_v)
        except: cls_v = np.nan
        try: chg_v = float(chg_v)
        except: chg_v = np.nan
        try: vol_raw = float(vol_raw)
        except: vol_raw = np.nan

        ry_top = y - ROW_H
        bg = LGREY if i % 2 == 0 else WHITE
        rrect(ax, x, ry_top, w, ROW_H * 0.88, bg, ec=MGREY, lw=0.2, r=0.03, zorder=1)
        ry = ry_top + ROW_H * 0.44

        # Rank circle
        rk = i + 1
        bc = accent if rk <= 3 else MGREY
        tc = WHITE  if rk <= 3 else DARK
        circle(ax, col["rk"], ry, 0.088, bc, zorder=2)
        ax.text(col["rk"], ry, str(rk),
                ha="center", va="center", fontsize=4.8,
                color=tc, fontweight="bold", zorder=3)

        # Symbol
        sym_txt = sym[:10] if len(sym) > 10 else sym
        ax.text(col["sym"]+0.04, ry, sym_txt,
                ha="left", va="center", fontsize=6.2,
                color=accent, fontweight="bold", zorder=3)

        # Close price
        ax.text(col["cls"], ry,
                f"{cls_v:,.1f}" if not np.isnan(cls_v) else "-",
                ha="center", va="center", fontsize=5.8,
                color=DARK, fontweight="bold", zorder=3)

        # Change % pill
        if not np.isnan(chg_v):
            cc  = GREEN if chg_v >= 0 else RED
            cbg = "#E9F7EF" if chg_v >= 0 else "#FDEDEC"
            rrect(ax, col["chg"]-0.27, ry-0.068, 0.54, 0.136, cbg, r=0.03, zorder=2)
            ax.text(col["chg"], ry,
                    f"{'+'if chg_v>=0 else ''}{chg_v:.2f}%",
                    ha="center", va="center", fontsize=5.2,
                    color=cc, fontweight="bold", zorder=3)
        else:
            ax.text(col["chg"], ry, "-", ha="center", va="center",
                    fontsize=5.2, color=MUTED, zorder=3)

        # Volume
        ax.text(col["vol"], ry, fmt_vol(vol_raw),
                ha="center", va="center", fontsize=5.0, color=DARK, zorder=3)

        y -= ROW_H

    return y


# ── Build combined single image ───────────────────────────────────────────────
def draw_combined_image(df, report_date, mc_map, out_path):
    # ── Prep data ─────────────────────────────────────────────────────────────
    bull_df = top10_trend(df, "Bullish")
    bear_df = top10_trend(df, "Bearish")
    neut_df = top10_trend(df, "Neutral")

    movers      = load_movers(report_date)
    active_df   = prep_mover_df(movers.get("most_active.csv", pd.DataFrame()), mc_map)

    bull_cnt  = len(df[df["Trend"].str.lower() == "bullish"])
    bear_cnt  = len(df[df["Trend"].str.lower() == "bearish"])
    neut_cnt  = len(df[df["Trend"].str.lower() == "neutral"])
    total_cnt = bull_cnt + bear_cnt + neut_cnt

    # ── Layout constants ──────────────────────────────────────────────────────
    FIG_W    = 7.2
    DPI      = 150
    MARGIN   = 0.18
    PANEL_W  = FIG_W - 2 * MARGIN
    COL_GAP  = 0.10
    COL2_W   = (PANEL_W - COL_GAP) / 2

    MAIN_HDR = 0.55
    PILL_H   = 0.25
    SEC_HDR  = 0.30
    COL_H    = 0.21
    ROW_H    = 0.225
    SEC_GAP  = 0.10
    FOOT_H   = 0.18

    def section_height(n=10):
        return SEC_HDR + COL_H + n * ROW_H + 0.04

    TOP_H = 0.10 + MAIN_HDR + 0.08 + PILL_H + 0.12
    FIG_H = TOP_H + section_height() + SEC_GAP + section_height() + 0.06 + FOOT_H + 0.10

    # ── Canvas ────────────────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(FIG_W, FIG_H))
    ax.set_xlim(0, FIG_W)
    ax.set_ylim(0, FIG_H)
    ax.axis("off")
    fig.patch.set_facecolor(BGPAGE)

    # ── Main header ───────────────────────────────────────────────────────────
    y = FIG_H - 0.10
    y -= MAIN_HDR
    rrect(ax, MARGIN, y, PANEL_W, MAIN_HDR, HEADER, r=0.08, zorder=3)
    ax.text(FIG_W/2, y + MAIN_HDR*0.68, "NSE DAILY MARKET REPORT",
            ha="center", va="center", fontsize=11, color=WHITE,
            fontweight="bold", zorder=5, family="monospace")
    ax.text(FIG_W/2, y + MAIN_HDR*0.26,
            f"Date: {report_date}   |   NSE Indicator Engine   |   Series: EQ",
            ha="center", va="center", fontsize=6.5, color=WHITE, alpha=0.85, zorder=5)

    # ── Summary pills ─────────────────────────────────────────────────────────
    y -= 0.08
    y -= PILL_H
    pill_w = PANEL_W / 4 - 0.05
    pill_colors = [GREEN, RED, BLUE, DARK]
    pill_labels = [f"Bullish\n{bull_cnt}", f"Bearish\n{bear_cnt}",
                   f"Neutral\n{neut_cnt}", f"Total EQ\n{total_cnt}"]
    for i, (col, lbl) in enumerate(zip(pill_colors, pill_labels)):
        px = MARGIN + i * (pill_w + 0.067)
        rrect(ax, px, y, pill_w, PILL_H, col, alpha=0.92, r=0.06, zorder=3)
        parts = lbl.split("\n")
        ax.text(px + pill_w/2, y + PILL_H*0.70, parts[0],
                ha="center", va="center", fontsize=5.5, color=WHITE,
                fontweight="bold", zorder=5)
        ax.text(px + pill_w/2, y + PILL_H*0.25, parts[1],
                ha="center", va="center", fontsize=7.5, color=WHITE,
                fontweight="bold", zorder=5)

    y -= 0.12

    LX = MARGIN
    RX = MARGIN + COL2_W + COL_GAP

    # ── Row 1: Bullish | Bearish ───────────────────────────────────────────────
    y1_start = y
    y1L = draw_section(ax, LX, y1_start, COL2_W,
                       "BULLISH TOP 10", GREEN, "^", bull_df,
                       SEC_HDR, COL_H, ROW_H)
    y1R = draw_section(ax, RX, y1_start, COL2_W,
                       "BEARISH TOP 10", RED,   "v", bear_df,
                       SEC_HDR, COL_H, ROW_H)
    y_after_row1 = min(y1L, y1R) - SEC_GAP

    # Thin divider
    ax.plot([MARGIN, MARGIN + PANEL_W],
            [y_after_row1 + SEC_GAP*0.5]*2,
            color=MGREY, lw=0.4, alpha=0.6)

    # ── Row 2: Neutral | Most Active ──────────────────────────────────────────
    y2L = draw_section(ax, LX, y_after_row1, COL2_W,
                       "NEUTRAL TOP 10", BLUE,   "~", neut_df,
                       SEC_HDR, COL_H, ROW_H)
    y2R = draw_section(ax, RX, y_after_row1, COL2_W,
                       "MOST ACTIVE",    ORANGE, "*", active_df,
                       SEC_HDR, COL_H, ROW_H)

    # ── Footer ────────────────────────────────────────────────────────────────
    fy = min(y2L, y2R) - 0.06
    ax.text(FIG_W/2, fy,
            f"Generated: {NOW.strftime('%d %b %Y  %I:%M %p IST')}"
            "   |   NSE Indicator Engine   |   Series: EQ",
            ha="center", va="top", fontsize=5.2, color=MUTED, zorder=3)

    plt.tight_layout(pad=0)
    fig.savefig(out_path, dpi=DPI, bbox_inches="tight",
                facecolor=BGPAGE, edgecolor="none")
    plt.close(fig)
    print(f"[IMAGE] Saved -> {out_path}")


# ── Send helpers ──────────────────────────────────────────────────────────────
def send_telegram(path, caption):
    if not TG_BOT_TOKEN or not TG_CHAT_ID:
        print("[TG] Credentials missing — skipping")
        return
    with open(path, "rb") as f:
        r = requests.post(
            f"https://api.telegram.org/bot{TG_BOT_TOKEN}/sendPhoto",
            data={"chat_id": TG_CHAT_ID, "caption": caption},
            files={"photo": f}, timeout=30)
    status = "OK" if r.status_code == 200 else f"FAIL {r.status_code}: {r.text[:120]}"
    print(f"[TG] {status}")


def _upload_imgbb(path):
    if not IMGBB_KEY:
        return None
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    r = requests.post("https://api.imgbb.com/1/upload",
                      data={"key": IMGBB_KEY, "image": b64}, timeout=30)
    if r.status_code == 200:
        return r.json()["data"]["url"]
    print(f"[ImgBB] Upload failed ({r.status_code}): {r.text[:100]}")
    return None


def send_whatsapp(path, caption):
    if not all([TWILIO_SID, TWILIO_TOKEN, WA_FROM]):
        print("[WA] Twilio credentials missing — skipping")
        return
    img_url = _upload_imgbb(path)
    if not img_url:
        print("[WA] No public image URL — skipping")
        return
    r = requests.post(
        f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_SID}/Messages.json",
        auth=(TWILIO_SID, TWILIO_TOKEN),
        data={"From": WA_FROM, "To": WA_TO, "Body": caption, "MediaUrl": img_url},
        timeout=30)
    status = "OK" if r.status_code in (200, 201) else f"FAIL {r.status_code}: {r.text[:120]}"
    print(f"[WA] {status}")


# ── Excel Trend Tracker ───────────────────────────────────────────────────────
def update_excel_tracker(df, report_date, bull_cnt, bear_cnt, neut_cnt, total_cnt):
    try:
        import openpyxl
        excel_path = "output_data/NSE_Trends_DateWise.xlsx"
        try:
            wb = openpyxl.load_workbook(excel_path)
        except Exception:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if "Daily_Summary" not in wb.sheetnames:
            ws_sum = wb.create_sheet("Daily_Summary")
            ws_sum.append(["Date", "Bullish", "Bearish", "Neutral",
                            "Total", "Top_Bullish", "Top_Bearish"])
            for col in ["A"]: ws_sum.column_dimensions[col].width = 14
            for col in ["B","C","D","E"]: ws_sum.column_dimensions[col].width = 10
            ws_sum.column_dimensions["F"].width = 40
            ws_sum.column_dimensions["G"].width = 40
        else:
            ws_sum = wb["Daily_Summary"]

        existing_dates = {str(row[0]) for row in ws_sum.iter_rows(min_row=2, values_only=True) if row[0]}

        if report_date not in existing_dates:
            top_bull = (df[df["Trend"].str.lower()=="bullish"]["NSE_SYMBOL"]
                        .value_counts().head(10).index.tolist()
                        if "Trend" in df.columns else [])
            top_bear = (df[df["Trend"].str.lower()=="bearish"]["NSE_SYMBOL"]
                        .value_counts().head(10).index.tolist()
                        if "Trend" in df.columns else [])
            ws_sum.append([report_date, bull_cnt, bear_cnt, neut_cnt, total_cnt,
                           ", ".join(top_bull), ", ".join(top_bear)])
            print(f"[EXCEL] Added {report_date} to Daily_Summary")

        sheet_name = report_date.replace("-","_")[:31]
        if sheet_name not in wb.sheetnames:
            ws_det = wb.create_sheet(sheet_name)
            cols = ["NSE_SYMBOL","Trend","Signal","ScannerName",
                    "CLOSE","PREV_CLOSE","CHANGE_PCT","VOLUME","ConfidenceScore"]
            cols = [c for c in cols if c in df.columns]
            ws_det.append(cols)
            for row in df[cols].fillna("").itertuples(index=False):
                ws_det.append(list(row))
            print(f"[EXCEL] Created sheet: {sheet_name} ({len(df)} rows)")

        wb.save(excel_path)
        print(f"[EXCEL] Saved -> {excel_path}")
    except Exception as e:
        print(f"[EXCEL] Error: {e}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("[START] NSE Image Report — Single Combined Image")
    df, report_date = load_report()

    if "Trend" not in df.columns:
        print("[ERROR] 'Trend' column not found in report")
        sys.exit(1)

    bull_cnt  = len(df[df["Trend"].str.lower() == "bullish"])
    bear_cnt  = len(df[df["Trend"].str.lower() == "bearish"])
    neut_cnt  = len(df[df["Trend"].str.lower() == "neutral"])
    total_cnt = bull_cnt + bear_cnt + neut_cnt
    print(f"[DATA] Bullish:{bull_cnt:,} | Bearish:{bear_cnt:,} | Neutral:{neut_cnt:,} | Total:{total_cnt:,}")

    mc_map = load_mc_mapping()
    dt     = report_date.replace("-", "_")
    out    = f"/tmp/NSE_DailyReport_{dt}.png"

    draw_combined_image(df, report_date, mc_map, out)

    caption = (
        f"*NSE Daily Market Report*\n"
        f"Date: {report_date}\n"
        f"Bullish: {bull_cnt:,} | Bearish: {bear_cnt:,} | Neutral: {neut_cnt:,}\n"
        f"Total EQ Signals: {total_cnt:,}"
    )

    send_telegram(out, caption)
    send_whatsapp(out, caption)
    update_excel_tracker(df, report_date, bull_cnt, bear_cnt, neut_cnt, total_cnt)

    print(f"[DONE] Single image sent | Telegram + WhatsApp | Excel updated")


if __name__ == "__main__":
    main()
